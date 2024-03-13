from os import path
from openpyxl import load_workbook

currentFilePath1 = path.abspath(__file__)
currentFilePath2 = currentFilePath1.replace('c:','C:')
currentFileName = 'variables.py'
PDVonTaskbar = path.abspath('imgs\PDVtaskBar.png')
FieldPassword = currentFilePath2.replace('variables.py','imgs\FieldPassword.png')
passPDV = '123456'

def importDadosExcel():
    arrayExcel= []
    ExcelcenarioDeTeste = path.join(currentPath,'cenariosDeTeste.xlsx')
    workbook = load_workbook(ExcelcenarioDeTeste)
    sheet = workbook['Venda'].iter_rows(min_row=2)

    for row in sheet:
        if row[0].value != 0:
            mountObjsRow = {'Produto':row[1].value,'Quantidade':row[2].value,'Nome':row[3].value, 'Id cenário':row[4].value,'O que':row[5].value, "Como":row[6].value, 'Objetivo':row[7].value}
            arrayExcel.append(mountObjsRow)

    return arrayExcel