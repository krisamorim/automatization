from os import path
from openpyxl import load_workbook

currentFilePath = path.abspath(__file__)
currentFileName = 'variables.py'
currentPath = currentFilePath.replace(currentFileName,"")
PDVonTaskbar = path.abspath('imgs\PDVtaskBar.png')
#login page varaibles
FieldPassword = currentFilePath.replace('variables.py','imgs\loginPage\FieldPassword.png')
FieldPasswordButton = FieldPassword.replace('FieldPassword','FieldPasswordButton')
listImgsLoginPage = [FieldPasswordButton, FieldPassword]
passPDV = '123456'
#sales screen variables
salesScreenCupom  = currentFilePath.replace('variables.py','imgs\salesScreen\salesScreenCupom.png')
listImgsSalesScreen = [salesScreenCupom]

def importDadosExcel():
    arrayExcel= []
    ExcelcenarioDeTeste = path.join(currentPath,'cenariosDeTeste.xlsx')
    workbook = load_workbook(ExcelcenarioDeTeste)
    sheet = workbook['Venda'].iter_rows(min_row=2)

    for row in sheet:
        if row[0].value != 0:
            mountObjsRow = {'Produto':row[1].value,'Quantidade':row[2].value,'Nome':row[3].value, 'Id cenário':row[4].value,'O que':row[5].value, "Como":row[6].value, 'Objetivo':row[7].value}
            arrayExcel.append(mountObjsRow)

    return arrayExcel